import requests
import time
import pyperclip
from tkinter import *
import openpyxl


def translate(content):  # use Google translation API
    key = 'AIzaSyBQYC3cdYM7NoP2Mx8OV_Z6-Z5VtRixPME'
    language_type = 'en'
    target = 'zh-TW'
    # content = 'translate'
    url = "https://translation.googleapis.com/language/translate/v2"
    data = {
        'key': key,                 # API key
        'source': language_type,    # input language
        'target': target,           # output language
        'q': content,
        'format': 'text'
    }
    headers = {'X-HTTP-Method-Override': 'GET'}
    # response = requests.post(url, data=data, headers=headers)
    response = requests.post(url, data)
    # print(response.json())
    # print(response)
    res = response.json()
    # print(res)
    result = res["data"]["translations"][0]["translatedText"]
    # print(result)
    return result


def getcopytext(copyBuff):  # get copied text
    copyedText = pyperclip.paste()
    if copyBuff != copyedText:
        copyBuff = copyedText
        normalizedText = copyBuff.replace('\r','\\r').replace('\n','\\n').replace('-\\r\\n','').replace("\\r\\n"," ")
    else:
        print('no change')
    return normalizedText


def showtxt(inputtxt):  # show the translation text
    root = Tk()
    root.title('translation')  # define the title
    root.geometry('200x100')  # the size of window
    label = Label(root, text=inputtxt, wraplength=400, justify='left', font=12)
    #label['text'] = inputtxt
    label.pack()
    root.mainloop()  # contain the window on the screen


def to_Excel(voc, translation):  # save the word to Excel
    wb = openpyxl.load_workbook('Vocabulary_list.xlsx')
    sheet = wb['sheet1']
    # print("最大行索引是：", sheet.max_row)
    data = sheet.values
    index = 1
    for i in data:
        if i[0] == voc:  # get the repeat word
            sheet['C'+str(index)] = i[2] + 1
            sheet['D'+str(index)] = None
            wb.save('Vocabulary_list.xlsx')
            wb.close()
            return 0

    maxRow = sheet.max_row  # get the maximum row
    x = maxRow + 1
    sheet['A' + str(x)] = voc
    sheet['B' + str(x)] = translation
    sheet['C' + str(x)] = 10
    wb.save('Vocabulary_list.xlsx')
    wb.close()
    return 0

if __name__ == "__main__":
    copyBuff = pyperclip.paste()
    while True:
        time.sleep(0.003)
        copyedText = pyperclip.paste()
        if copyBuff != copyedText:
            copyBuff = copyedText
            voc = copyBuff.replace('\r', '\\r').replace('\n', '\\n').replace('-\\r\\n', '').replace("\\r\\n", " ")
            translation = translate(voc)
            showtxt(translation)
            to_Excel(voc, translation)
            print('save success')