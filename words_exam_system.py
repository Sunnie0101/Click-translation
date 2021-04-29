import openpyxl
import datetime
import random
import os

"""
the introduce of the column and the rules are in the report
"""


def new_voc_list(filename, number):  # create today's vocabulary list
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet['A1'] = 'voc'
    sheet['B1'] = 'translation'
    sheet['C1'] = 'mode'
    sheet['D1'] = 'situation'
    sheet['E1'] = 'value1'
    sheet['F1'] = 'value2'

    voc_test, translation_test = choice_voc(number)
    x = 2
    for i in range(len(voc_test)):
        sheet['A' + str(x)] = voc_test[i]
        sheet['B' + str(x)] = translation_test[i]
        sheet['C' + str(x)] = 0
        sheet['D' + str(x)] = 0
        sheet['F' + str(x)] = 0
        x += 1

    wb.save(filename)
    return voc_test, translation_test


def choice_voc(number):  # choose today's vocabularies
    wb_data = openpyxl.load_workbook('Vocabulary_list.xlsx')
    sheet_data = wb_data['sheet1']

    voc_test = []           # vocabulary for test
    translation_test = []   # translation for test

    voc_a = []
    translation_a = []

    data = sheet_data.values
    # print(data)
    for i in data:
        # print(i)
        if i[3] is None or i[3] == 'wrong':
            voc_test.append(i[0])
            translation_test.append(i[1])
    # print(len(voc))
    # print(translation)

    if len(voc_test) > number:  # the collected voc over the need
        result_list = random.sample(range(0, len(voc_test) - 1), len(voc_test) - number)
        # print('res1:', result_list)
        result_list.sort(reverse=True)
        for j in range(len(result_list)):
            del voc_test[result_list[j]]
            del translation_test[result_list[j]]
        # print('res2:', result_list)

    elif len(voc_test) < number:  # the collected voc lower the need
        data = sheet_data.values
        for i in data:
            # print(i)
            if i[3] == 0:
                voc_a.append(i[0])
                translation_a.append(i[1])
        for j in range(len(voc_a)):
            k = random.randint(0, len(voc_a) - 1)
            voc_test.append(voc_a.pop(k))
            translation_test.append((translation_a.pop(k)))
            if voc_a == [] or len(voc_test) == number:  # the sum of the vocabularies is enough
                break
    # print(voc)
    # print(translation)
    wb_data.close()

    return voc_test, translation_test


def exam(filename, voc_all, translation_all):  # exam system
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Sheet']

    sum = len(voc_all)
    while sum != 0:
        k = random.randint(0, len(voc_all)-1)
        print('k:', k)
        if sheet['C' + str(k+2)].value == 0:

            print(voc_all)
            print(translation_all)
            voc_exam = voc_all[k]
            translation_exam = translation_all[k]
            res = option_exam(voc_exam, translation_exam, translation_all)

            print(sheet['F' + str(k + 2)].value)
            sheet['F' + str(k + 2)] = sheet['F' + str(k + 2)].value + 1
            if res is True:
                sheet['C' + str(k + 2)] = 1
            elif res is False:
                sheet['E' + str(k + 2)] = 'wrong'

        elif sheet['C' + str(k+2)].value == 1:
            print(voc_all)
            print(translation_all)
            res = fill_exam(voc_all[k], translation_all[k])
            sheet['F' + str(k + 2)] = sheet['F' + str(k + 2)].value + 1

            if res is True:
                sheet['D' + str(k + 2)] = sheet['D' + str(k + 2)].value + 1
                if sheet['D' + str(k + 2)].value == 1:
                    sheet['C' + str(k + 2)] = 2
                    sum -= 1
            elif res is False:
                sheet['D' + str(k + 2)] = sheet['D' + str(k + 2)].value - 1
                if sheet['D' + str(k + 2)].value == -2:
                    sheet['C' + str(k + 2)] = 0

        wb.save(filename)

    return 0


def option_exam(voc, translation, translations_all):  # mode1:multiple-choice question (en->zh)
    options = translations_all.copy()
    options.remove(translation)
    options = random.sample(options, 2)
    k = random.randint(0, 2)
    options.insert(k, translation)
    print('題目:', voc)
    print('1){} 2){} 3){}'.format(options[0], options[1], options[2]))

    ans = eval(input('your answer:'))
    if ans == k+1:
        print("Correct!")
        return True
    else:
        print("The answer is", translation)
        return False


def fill_exam(voc, translation):  # mode2:filling question (zh->en)
    print('題目:', translation)
    ans = input('your answer:')
    if ans == voc:
        print("Correct!")
        return True
    else:
        while(ans != voc):
            print("The answer is", voc)
            print('Try again!')
            print('題目:', translation)
            ans = input('your answer:')
            print('your ans:', ans)
        print("Correct!")
        return False


def back_to_data(filename, voc_test, translation_test):
    wb_data = openpyxl.load_workbook('Vocabulary_list.xlsx')
    sheet_data = wb_data['sheet1']
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Sheet']
    data = sheet_data.values
    index = 1
    times = 0
    for i in data:
        if index != 1:
            # print(i[2])
            if i[2] <= 5:
                value1 = -1
                value2 = 0
                value3 = 0
            elif i[2] >= 11:
                value1 = -3
                value2 = -1
                value3 = 0
            else:
                value1 = -2
                value2 = -1
                value3 = 1
            for j in range(2, len(voc_test) + 1):
                if voc_test[j-2] == i[0]:
                    times += 1
                    if sheet['F'+str(j)].value == 2:
                        sheet_data['C'+str(index)] = sheet_data['C'+str(index)].value + value1
                        sheet_data['D' + str(index)] = 0
                    elif sheet['E'+str(j)].value == 'wrong':
                        print('yes')
                        sheet_data['C'+str(index)] = sheet_data['C'+str(index)].value + value3
                        sheet_data['D'+str(index)] = 'wrong'
                    else:
                        sheet_data['C' + str(index)] = sheet_data['C' + str(index)].value + value2
                        sheet_data['D' + str(index)] = 0
                    wb_data.save('Vocabulary_list.xlsx')
                    break
        if times == len(voc_test):
            break
        index += 1
    return 0


def get_voc(filename):  # get today's vocabularies
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Sheet']
    data = sheet.values
    voc_test = []
    translation_test = []
    index = -1
    for i in data:
        if index >= 0:
            voc_test.append(i[0])
            translation_test.append(i[1])
        index += 1
    return voc_test, translation_test


if __name__ == '__main__':
    today = datetime.date.today()
    today = str(today)                      # ex:2020-06-18
    file = "Vocabulary" + today + ".xlsx"   # ex:Today's vocabulary2020-06-18.xlsx
    ever = os.path.isfile(file)             # whether the file exist or not
    if ever:
        print('start practice again~')
        voc, translation = get_voc(file)
    else:
        num = int(input('how many vocabularies you want to recite today?'))
        print('start practice~')
        voc, translation = new_voc_list(file, num)

    exam(file, voc, translation)
    print('\nfinish!')

    print('loading')
    back_to_data(file, voc, translation)
    print('Done!')
